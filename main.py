import openpyxl as ox
from openpyxl import load_workbook

from analisis import *
from api_client import get_weather, get_ciudad, get_pais, get_coordenadas, get_weatherDetail
import pandas as pd


from translate import Translator


traductorIngles = Translator(to_lang="en")
traductorEspañol = Translator(to_lang="es")


def dia_o_noche(parametro):

    if parametro == 1:
        actual = "dia"
    else:
        actual = "noche"
    return actual

def direccion_viento(grados):
    direccion = None
    if (grados >= 337.5 or grados < 22.5):
        direccion = "Norte"
    elif grados < 67.5:
        direccion = "Noreste"
    elif grados < 112.5:
        direccion = "Este"
    elif grados < 157.5:
        direccion = "Sureste"
    elif grados < 202.5:
        direccion = "Sur"
    elif grados < 247.5:
        direccion = "Suroeste"
    elif grados < 292.5:
        direccion = "Oeste"
    else:
        direccion = "Noroeste"

    return direccion


def estadoTiempo(parametro):
    estado = None
    if parametro == 0:
        estado = "Despejado"
    elif parametro == 1:
        estado = "Parcialmente nublado"
    elif parametro == 2:
        estado = "Nublado"
    elif parametro == 3:
        estado = "Muy nublado"

    return estado


def add_rows(filas):
    workbook = ox.load_workbook("clima.xlsx")

    clima_sheet = workbook["clima"]
    numFila = clima_sheet.max_row + 1
    numFila = str(numFila)
    letrasColumna = ["A", "B", "C", "D", "E", "F", "G", "H"]

    for i in range(len(letrasColumna)):
        celdaFinal = str(letrasColumna[i] + numFila)

        clima_sheet[celdaFinal] = filas[i]
    workbook.save("clima.xlsx")
    workbook.close()


def mostrarMenu():
    print("----MENÚ-----")
    print("1-- Mostar tiempos climáticos actuales")
    print("2-- Añadir tiempos climáticos  ")
    print("3-- Eliminar tiempos climáticos ")
    print("4-- Mostrar altitud ")
    print("5-- Filtrar tiempo  climátco por ciudad")
    print("6-- Filtrar tiempo climático por país")
    print("7-- Mostrar mayor temperatura")
    print("8-- Mostrar menor temperatura")
    print("9-- Mostrar Velocidad del viento")
    print("10- Salir")

def mostrarTemperaturaActual():
    pais = input("Introduce un país: ")
    pais = pais.capitalize()

    pais = traductorIngles.translate(pais)
    pais = pais.lower()
    pais = pais.strip()
    print(pais)

    while not get_pais(pais):
        print("El país introducido  no existe ")
        pais = input("Por favor introduzca un país correcto: ")
        pais = pais.strip()

    ciudad = input("Introduce una ciudad: ")
    ciudad = ciudad.capitalize()
    ciudad = ciudad.strip()
    print(ciudad)

    while not get_ciudad(ciudad):
        print("La ciudad  introducida no existe ")
        ciudad = input("Por favor introduzca una ciudad correcta: ")
        ciudad = ciudad.strip()

    latitud, longitud = get_coordenadas(ciudad, pais)
    print(latitud, longitud)

    print(f"***************TIEMPO EN {ciudad.upper()}******************")
    weather = get_weather(latitud, longitud)
    weather_detail = get_weatherDetail(latitud, longitud)

    print("Temperatura :", weather["temperature"], "Cº")

    print("Altitud sobre el nivel del mar: ", weather_detail["elevation"], "metros")

    print("Viento:", weather["windspeed"], "km/h")

    print("Actualmente es de ", dia_o_noche(weather["is_day"]))

    print("Dirección del viento: ", direccion_viento(weather["winddirection"]))

    print("Estado del clima: ", estadoTiempo(weather["weathercode"]))

    print("\nRespuesta generada en: ", weather_detail["generationtime_ms"], "milisegundos ")


def añadirTiempoClimatico():

    filaDatos = []
    pais = input("Introduce un país: ")

    pais = traductorIngles.translate(pais)
    pais = pais.strip()
    pais = pais.lower()
    print(pais)

    while not get_pais(pais):
        print("El país introducido  no existe ")
        pais = input("Por favor introduzca un país correcto: ")
        pais = pais.lower()
        pais = pais.strip()
        pais = traductorIngles.translate(pais)

    filaDatos.append(traductorEspañol.translate(pais.capitalize()))  # Añadimos el país

    ciudad = input("Introduce una ciudad: ")
    ciudad = ciudad.lower()
    ciudad = ciudad.strip()


    while not get_ciudad(ciudad):
        print("La ciudad  introducida no existe ")
        ciudad = input("Por favor introduzca una ciudad correcta: ")
        ciudad = ciudad.lower()
        ciudad = ciudad.strip()





    filaDatos.append((ciudad.capitalize()))  # Añadimos ciudad

    latitud, longitud = get_coordenadas(ciudad, pais)
    if latitud is None or longitud is None:
        print("No se encontraron coordenadas para esa ciudad.")
        return

    weather = get_weather(latitud, longitud)

    if weather is None:
        print("No se pudo obtener el clima.")
        return
    
    weather_detail = get_weatherDetail(latitud, longitud)

    filaDatos.append(weather["temperature"])

    filaDatos.append(weather_detail["elevation"])  # añadimos altiud

    filaDatos.append(weather["windspeed"])  # añadimos velocidad del viento

    filaDatos.append(dia_o_noche(weather["is_day"]))  # añadimos si es de día o no

    filaDatos.append(direccion_viento(weather["winddirection"]))  # añadimos la dirección del viento

    filaDatos.append(estadoTiempo(weather["weathercode"]))  # Añadimos el estado del tiempo


    if comprobarSiExiste(pais,ciudad):
        print("----------------------------EL TIEMPO CLIMÁTICO REGISTRADO YA EXISTE-------------------")

    else:

      add_rows(filaDatos)
      print("-------------------TIEMPO CLIMÁTICO REGISTRADO CORRECTAMENTE--------------")




def eliminarTiempoClimático():
    paisEliminado = input("Introduzca el país registrado: ").strip().lower()
    ciudadEliminada = input("Introduzca la ciudad eliminada: ").strip().lower()

    archivo = "clima.xlsx"
    workbook = load_workbook(archivo)
    clima_sheet = workbook["clima"]

    eliminacionPosible = False

    for i in range(clima_sheet.max_row, 1, -1):

        pais = clima_sheet[f"A{i}"].value
        ciudad = clima_sheet[f"B{i}"].value

        pais_excel = str(pais).strip().lower()
        ciudad_excel = str(ciudad).strip().lower()

        if pais_excel == paisEliminado and ciudad_excel == ciudadEliminada:
            clima_sheet.delete_rows(i)
            eliminacionPosible = True

    if eliminacionPosible:
        workbook.save(archivo)
        print("---------------------REGISTRO ELIMINADO CORRECTAMENTE✅---------------------")
    else:
        print("-------------------NO SE HA PODID HACER EL REGISTRO ❌-------------------")
        print("No se encontró el registro ❌")

    workbook.close()

def comprobarSiExiste(pais,ciudad):

    archivo = "clima.xlsx"
    workbook = load_workbook(archivo)
    clima_sheet = workbook["clima"]

    encontrado = False

    for i in range(clima_sheet.max_row, 1, -1):

        pais = clima_sheet[f"A{i}"].value
        ciudad = clima_sheet[f"B{i}"].value

        pais_excel = str(pais).strip().lower()
        ciudad_excel = str(ciudad).strip().lower()

        if pais_excel.capitalize() == pais.capitalize() and ciudad.capitalize() == ciudad_excel.capitalize():
            encontrado = True

        return encontrado

salir = False

while not salir:

        mostrarMenu()
        opcionMenu = input("Introduzca una opción del menú (1-10): ")

        opcionMenu = int(opcionMenu)

        if opcionMenu == 1:
            mostrar_toda_tabla()
        elif opcionMenu == 2:
            añadirTiempoClimatico()
        elif opcionMenu == 3:
            eliminarTiempoClimático()
        elif opcionMenu == 4:
            get_altitud()
        elif opcionMenu == 5:
            filtrar_por_ciudad()
        elif opcionMenu == 6:
            filtrar_por_pais()
        elif opcionMenu == 7:
            mostrar_mayor_temperatura()
        elif opcionMenu == 8:
            mostrar_menor_temperatura()
        elif opcionMenu == 9:
            mostrar_velocidad_viento()
        elif opcionMenu == 10:
            print("Fin del programa")
            salir = True

        else:
            opcionMenu = input("La opción escogida es incorecta\nPor favor escoga una de las opciones del menú(1-10): ")

